"""
course_audit_tool.py
====================
GUI application for auditing standardised EEE course folders at UCT.

Produces two output artefacts saved into the scanned course root:
  - A plain-text audit log  (<date>_<course>_folder_audit.txt)
  - An Excel review workbook (<date>_<course>_folder_audit.xlsx)

Supported folder-structure profiles
-------------------------------------
  Legacy   — 2023/2024 layout: dot-numbered, (h) suffixes, 09. Exam,
             13. Supplementary Exam, 11. Additional resources with subfolders
  Current  — 2025 layout: dot-numbered, no (h), 12/13 Admin exam folders,
             10. Additional resources, 11. Other
  New      — 2026 layout: lower_snake_case, 8 folders, 08_exams two-level

GA moderation
--------------
  GA courses require a compulsory 00_ga_moderation folder (or legacy
  equivalent). Detected automatically from the course code. The
  assessment forms subfolder is validated against the DP list student count.

Auto-detection
---------------
  Year is extracted from the folder name first; the disk structure is used
  as a fallback. GA is detected from the course code list.

NONE convention
----------------
  Any folder containing the word NONE (any position, any separator) is
  treated as intentionally empty.
    - Empty NONE folder      → NONE - ACCEPTED
    - NONE folder with files → POPULATED DESPITE NONE

Matching strategy
------------------
  Both template and disk names are normalised identically:
    1. Strip NONE markers
    2. Strip admin status markers (MISSING, INCOMPLETE, URGENT, etc.)
    3. Strip (h) and other parenthetical suffixes
    4. Strip leading number/letter prefix (09., a., 01_, a_)
    5. Normalise separators (space / hyphen / underscore all equivalent)
    6. Lowercase
    7. Strip trailing s
  So "c. GA assessment forms (for students)", "c_ga_assessment_forms",
  and "c. GA Assessment Forms" all resolve to "ga assessment form".
"""

# ---------------------------------------------------------------------------
# Standard-library imports
# ---------------------------------------------------------------------------
import csv
import json
import os
import re
import sys
import traceback
from collections import Counter
from datetime import datetime
from typing import ClassVar

# ---------------------------------------------------------------------------
# GUI toolkit
# ---------------------------------------------------------------------------
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk

# ---------------------------------------------------------------------------
# Excel workbook generation
# ---------------------------------------------------------------------------
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

# ---------------------------------------------------------------------------
# Image handling (Pillow — optional)
# ---------------------------------------------------------------------------
try:
    from PIL import Image, ImageTk  # type: ignore
    _PILLOW_AVAILABLE = True
except ImportError:
    _PILLOW_AVAILABLE = False


# ===========================================================================
# Global constants
# ===========================================================================

IGNORED_FILES: set[str] = {".ds_store"}

# Pattern to detect files whose stem ends with "unsigned"
_UNSIGNED_PATTERN = re.compile(r'[\s\-_]unsigned$', re.IGNORECASE)

# Authorised auditors shown in the dropdown
AUDIT_USERS: list[str] = [
    "Verrinder",
    "YAbdul Gaffar",
    "Mwangama",
    "Buxey",
    "Harris",
    "Langenhoven",
]

# Teaching assistants shown in the TA dropdown (editable — type to add a custom name)
TA_NAMES: list[str] = [
    "",
    "Obbo",
    "Moradi",
    "Islam",
    "Ferreiro",
    "Tallack",
    "Mwangi",
    "Tetteh",
    "Kalonji",
    "Maqungu",
    "Mahlaola",
    "Magavha",
    "Maimbo",
    "Makhubela",
    "Webb",
]

# GA course codes — case-insensitive, year-independent
GA_COURSES: set[str] = {
    "eee4022s", "eee4022f",
    "eee3096s", "eee3100s",
    "eee4118f", "eee4119f", "eee4120f", "eee4121f", "eee4126f",
    "eee4124c", "eee4125c", "eee4113f", "eee3088f",
    "eee3097s", "eee3098s", "eee3099s",
}

# DP list column names (lowercased) that identify the student-number column
DP_LIST_STUDENT_COLUMNS: set[str] = {
    # UCT PeopleSoft export columns
    "emplid", "campus id",
    # Generic variants
    "student number", "student_number", "studentnumber",
    "student no", "student no.", "student id",
}

# Keys identifying the GA moderation folder regardless of naming style
_GA_MOD_KEYS: set[str] = {"00_ga_moderation", "ga moderation folder", "ga moderation"}

# Admin status markers that may be appended to folder names
_STATUS_MARKER_RE = re.compile(
    r'(?<=\w)[\s\-_]+(MISSING|INCOMPLETE|URGENT|TODO|UNSIGNED|EMPTY|TO\s+BE\s+SIGNED)[\s\-_\S]*$',
    re.IGNORECASE,
)


# ===========================================================================
# Profile definitions
# ===========================================================================

LEGACY_STRUCTURE: dict[str, list[str]] = {
    "01. Administration": [
        "a. Course handouts",
        "b. Prescribed texts",
        "c. Course evaluations",
        "d. DP list",
    ],
    "02. Notes": [],
    "03. Lessons": [
        "a. Slides",
        "b. Recordings",
        "c. Additional material",
    ],
    "04. Tutorials": [
        "a. Instruction sheets",
        "b. Recordings",
        "c. Solutions",
        "d. Sample hand-ins",
    ],
    "05. Practicals": [
        "a. Instruction sheets",
        "b. Recordings",
        "c. Solutions",
        "d. Sample hand-ins",
    ],
    "06. Assignments": [
        "a. Instruction sheets",
        "b. Solutions",
        "c. Sample hand-ins",
    ],
    "07. Projects": [
        "a. Instruction sheets",
        "b. Sample hand-ins",
    ],
    "08. Tests": [
        "a. Questions",
        "b. Model answers",
        "c. Sample answers from students",
    ],
    "09. Exam": [
        "a. Exam paper",
        "b. Exam model answer",
        "c. Exam scripts",
        "d. External moderator reports",
        "e. Departmental control sheet",
        "f. Mark sheets",
    ],
    "10. Software": [],
    "11. Additional resources": [
        "a. Past tests and exams",
        "b. datasheets",
        "c. code",
    ],
    "12. Other": [
        "Images",
        "Sick notes",
    ],
    "13. Supplementary Exam": [
        "a. Exam paper",
        "b. Exam model answer",
        "c. Exam scripts",
        "d. External moderator reports",
        "e. Mark sheets",
    ],
}

CURRENT_STRUCTURE: dict[str, list[str]] = {
    "01. Administration": [
        "a. Course handouts",
        "b. Prescribed texts",
        "c. Course evaluations",
        "d. DP list final",
    ],
    "02. Notes": [],
    "03. Lessons": [
        "Slides",
        "Recordings",
        "Additional material",
    ],
    "04. Tutorials": [
        "a. Instruction sheets",
        "b. Solutions",
        "c. Sample hand-ins",
    ],
    "05. Practicals": [
        "a. Instruction sheets",
        "b. Solutions",
        "c. Sample hand-ins",
    ],
    "06. Assignments": [
        "a. Instruction sheets",
        "b. Solutions",
        "c. Sample hand-ins",
    ],
    "07. Projects": [
        "a. Instruction sheets",
        "b. Sample hand-ins",
    ],
    "08. Tests": [
        "a. Questions",
        "b. Model answers",
        "c. Sample answers from students",
    ],
    "09. Software": [],
    "10. Additional resources": [
        "a. Past tests and exams",
        "b. datasheets",
        "c. code",
    ],
    "11. Other": [],
    "12. Exams Main (Admin)": [
        "a. Exam paper",
        "b. Exam model answer",
        "c. External moderator reports",
        "d. Departmental control sheet",
        "e. Exam scripts",
        "f. Mark sheets",
    ],
    "13. Exams SUPPS (Admin)": [
        "a. Exam paper",
        "b. Exam model answer",
        "c. External moderator reports",
        "d. Departmental control sheet",
        "e. Exam scripts",
        "f. Mark sheets",
    ],
}

NEW_STRUCTURE: dict[str, list[str]] = {
    "01_administration": [
        "a_course_handout",
        "b_dp_list",
        "c_course_evaluation",
        "d_marksheets",
    ],
    "02_teaching_materials": [
        "a_prescribed_textbooks",
        "b_notes",
        "c_slides",
    ],
    "03_tutorials": [
        "a_handouts",
        "b_solutions",
        "c_hand-ins",
    ],
    "04_practicals": [
        "a_handouts",
        "b_solutions",
        "c_hand-ins",
    ],
    "05_assignments": [
        "a_handouts",
        "b_solutions",
        "c_hand-ins",
    ],
    "06_projects": [
        "a_handouts",
        "b_solutions",
        "c_hand-ins",
    ],
    "07_tests": [
        "a_papers",
        "b_solutions",
        "c_scripts",
    ],
    "08_exams": [
        "01_main",
        "02_supp_de",
    ],
}

# ---------------------------------------------------------------------------
# EEE4022 structure  (project-based GA course — no standard profile folders)
# ---------------------------------------------------------------------------
# EEE4022S/F is a capstone project course with its own folder convention.
# Detection: course code EEE4022 anywhere in the folder name.
EEE4022_STRUCTURE: dict[str, list[str]] = {
    "1. Course Handout": [],
    "2. DP List":        [],
    "3. Final Results":  [],
    "4. Student Projects":           [],   # validated separately per-student
    "5. Moderators":                 [],   # validated separately per-moderator
    "6. GA External Moderators Report": [],
}

# Subfolders inside each exam group for the New profile

NEW_EXAM_SUBFOLDERS: list[str] = [
    "a_papers",
    "b_solutions",
    "c_scripts",
    "d_external_moderator_reports",
]

# New-style GA moderation subfolders (2026 snake_case)
NEW_GA_MOD_SUBFOLDERS: list[str] = [
    "a_dp_tracking",
    "b_assessment_instructions",
    "c_assessment_forms",
    "d_ca_results",
    "e_evidence_sorted_by_students_alphabetically",
    "f_external_moderator_report",
]

# Legacy GA moderation subfolders (pre-2026)
LEGACY_GA_MOD_SUBFOLDERS: list[str] = [
    "a. Submission tracking",
    "b. Instructions to students regarding GA assessments",
    "c. GA assessment forms (for students)",
    "d. Continuous assessment results",
    "e. Evidence sorted by Students Alphabetically",
    "f. DP list",
    "g. External Moderator Report (GA)",
]

STRUCTURE_PROFILES: dict[str, dict[str, list[str]]] = {
    "Legacy":  LEGACY_STRUCTURE,
    "Current": CURRENT_STRUCTURE,
    "New":     NEW_STRUCTURE,
    "EEE4022": EEE4022_STRUCTURE,
}

# Flat set of all normalised base names across every profile (for UNEXPECTED hints)
_ALL_TEMPLATE_BASE_NAMES: set[str] = set()


# ===========================================================================
# Module-level normalisation helpers
# ===========================================================================

def _strip_none(name: str) -> str:
    """Remove NONE token and surrounding separators from *name*."""
    return re.sub(
        r'[\s\-_]*(?<![A-Za-z])NONE(?![A-Za-z])[\s\-_]*',
        ' ', name.strip(), flags=re.IGNORECASE,
    ).strip()


def _strip_admin_markers(name: str) -> str:
    """Remove trailing administrator status-marker annotations."""
    name = name.strip()
    marker = r'(MISSING|INCOMPLETE|URGENT|TODO|UNSIGNED|EMPTY|TO\s+BE\s+SIGNED)'
    # Try compound strip: uppercase qualifier + marker
    compound = re.sub(
        rf'([\s\-_]+[A-Z]{{2,}}[\s\-_]+{marker})[\s\-_\S]*$', '', name,
    ).strip()
    base_words = re.sub(r'^[0-9a-z]+[._\-]\s*', '', compound, flags=re.IGNORECASE).split()
    if len(base_words) >= 2 and compound != name:
        return compound
    return re.sub(
        rf'(?<=\w)[\s\-_]+{marker}[\s\-_\S]*$', '', name, flags=re.IGNORECASE,
    ).strip()


def _strip_parentheticals(name: str) -> str:
    """Remove trailing parenthetical suffixes like (h), (admin), (CHECK)."""
    return re.sub(r'\s*\([^)]*\)\s*$', '', name).strip()


def _strip_prefix(name: str) -> str:
    """Remove leading alphanumeric prefix and separator (09., a., 01_, a_)."""
    # Handles both dot style "09. " and underscore style "01_"
    return re.sub(r'^[0-9a-z]+[._\-]\s*', '', name.strip(), flags=re.IGNORECASE).strip()


def _normalise_separators(name: str) -> str:
    """Replace hyphens and underscores with spaces, collapse multiple spaces."""
    return re.sub(r'[\-_]+', ' ', name).strip()


def nbk(name: str) -> str:
    """Return the normalised base key for *name*.

    Applied uniformly to both template names and disk names so that any
    reasonable variation of the same folder concept resolves to the same key.

    Pipeline:
      strip NONE → strip admin markers → strip parentheticals →
      strip prefix → normalise separators → lowercase → strip trailing s
    """
    key = _strip_prefix(
        _strip_parentheticals(
            _strip_admin_markers(
                _strip_none(name)
            )
        )
    )
    key = _normalise_separators(key).lower()
    if key.endswith('s') and len(key) > 2:
        key = key[:-1]
    return key


def _has_none(name: str) -> bool:
    """Return True if *name* contains the word NONE as a discrete token."""
    return bool(re.search(r'(?<![A-Za-z])NONE(?![A-Za-z])', name.strip(), re.IGNORECASE))


def _has_admin_marker(name: str) -> bool:
    """Return True if *name* ends with a recognised admin status marker."""
    return bool(_STATUS_MARKER_RE.search(name))


# Populate _ALL_TEMPLATE_BASE_NAMES after nbk is defined
for _profile in STRUCTURE_PROFILES.values():
    for _name, _children in _profile.items():
        _ALL_TEMPLATE_BASE_NAMES.add(nbk(_name))
        for _child in _children:
            _ALL_TEMPLATE_BASE_NAMES.add(nbk(_child))
for _sub in NEW_GA_MOD_SUBFOLDERS + LEGACY_GA_MOD_SUBFOLDERS:
    _ALL_TEMPLATE_BASE_NAMES.add(nbk(_sub))


# ===========================================================================
# Main application class
# ===========================================================================

class CourseFolderAuditApp:
    """Tkinter GUI for auditing EEE course folder structures at UCT."""

    # ------------------------------------------------------------------ #
    #  GUI colour palette (EEE department branding)                        #
    # ------------------------------------------------------------------ #
    C_CHROME  = "#BFCCC2"  # EEE pale green — window chrome
    C_CONTENT = "#FFFFFF"  # White — content areas
    C_INK     = "#6C9273"  # EEE mid green — buttons, headings
    C_TEXT    = "#000000"  # Black — all text
    C_UCT     = "#003C69"  # UCT dark blue — Excel headers only

    # ------------------------------------------------------------------ #
    #  Status → (background_hex, font_hex)                                 #
    # ------------------------------------------------------------------ #
    STATUS_COLOURS: ClassVar[dict[str, tuple[str, str]]] = {
        "OK":                     ("B6D7A8", "000000"),
        "EMPTY - REVIEW":         ("FFE599", "000000"),
        "MISSING":                ("EA9999", "000000"),
        "MISSING CHILDREN":       ("EA9999", "000000"),
        "UNEXPECTED":             ("F9CB9C", "000000"),
        "NONE - ACCEPTED":        ("9FC5E8", "000000"),
        "POPULATED DESPITE NONE": ("B4A7D6", "000000"),
        "REVIEW - HAND-INS":      ("F9CB9C", "000000"),
        "DUPLICATE":              ("EA9999", "000000"),
        "ADMIN FLAG":             ("FFD966", "000000"),
        "REVIEW - GA INCOMPLETE": ("F4CCCC", "000000"),
        "UNSIGNED FILE":          ("FCE4D6", "000000"),
    }

    # Statuses that appear in the Issues tab / Exceptions sheet
    _ISSUE_STATUSES: ClassVar[set[str]] = {
        "MISSING", "EMPTY - REVIEW", "UNEXPECTED",
        "POPULATED DESPITE NONE", "REVIEW - HAND-INS",
        "DUPLICATE", "ADMIN FLAG", "REVIEW - GA INCOMPLETE",
        "UNSIGNED FILE",
    }

    # ==================================================================== #
    #  Initialisation                                                        #
    # ==================================================================== #

    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("Course Audit Tool")
        self.root.geometry("1420x860")
        self.root.minsize(1120, 720)
        self.root.configure(bg=self.C_CHROME)

        self.selected_directory = tk.StringVar()
        self.profile_mode       = tk.StringVar(value="Auto-detect")
        self.selected_user      = tk.StringVar(value=AUDIT_USERS[0])
        self.selected_ta        = tk.StringVar(value="")

        self.recent_dirs_file = os.path.join(
            os.path.expanduser("~"), ".course_folder_audit_recent.json"
        )
        self.recent_directories = self._load_recent_dirs()

        self._ref_student_count: int | None = None  # set per audit run
        self._img_uct = self._load_logo("logo_uct.png", 36)
        self._img_eee = self._load_logo("logo_eee.png", 36)
        if self._img_uct:
            self.root.iconphoto(True, self._img_uct)

        self._apply_theme()
        self._build_gui()

    # ==================================================================== #
    #  Logo loading                                                          #
    # ==================================================================== #

    def _load_logo(self, filename: str, height: int):
        if not _PILLOW_AVAILABLE:
            return None
        base = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
        path = os.path.join(base, filename)
        if not os.path.isfile(path):
            return None
        try:
            img   = Image.open(path).convert("RGBA")
            ratio = height / img.height
            img   = img.resize((max(1, int(img.width * ratio)), height), Image.LANCZOS)
            return ImageTk.PhotoImage(img)
        except Exception:
            return None

    # ==================================================================== #
    #  Theme                                                                 #
    # ==================================================================== #

    def _apply_theme(self) -> None:
        style = ttk.Style(self.root)
        style.theme_use("clam")

        style.configure("TNotebook", background=self.C_CHROME, borderwidth=0)
        style.configure("TNotebook.Tab", background=self.C_CHROME, foreground="#000000",
                        padding=[12, 6], font=("Montserrat", 10, "bold"))
        style.map("TNotebook.Tab",
                  background=[("selected", self.C_CHROME)],
                  foreground=[("selected", "#000000")])

        style.configure("Treeview", background=self.C_CONTENT, foreground=self.C_TEXT,
                        fieldbackground=self.C_CONTENT, rowheight=24,
                        font=("Montserrat", 10))
        style.configure("Treeview.Heading", background=self.C_INK, foreground="#FFFFFF",
                        font=("Montserrat", 10, "bold"), relief="flat")
        style.map("Treeview",
                  background=[("selected", self.C_INK)],
                  foreground=[("selected", self.C_TEXT)])
        style.map("Treeview.Heading", background=[("active", self.C_INK)])

        style.configure("TCombobox", fieldbackground=self.C_CONTENT,
                        background=self.C_CONTENT, foreground=self.C_TEXT,
                        arrowcolor=self.C_TEXT, selectbackground=self.C_INK,
                        selectforeground=self.C_TEXT)
        style.map("TCombobox",
                  fieldbackground=[("readonly", self.C_CONTENT)],
                  foreground=[("readonly", self.C_TEXT)])

        style.configure("TScrollbar", background=self.C_INK,
                        troughcolor=self.C_CHROME, arrowcolor=self.C_TEXT,
                        bordercolor=self.C_CHROME)

    # ==================================================================== #
    #  GUI construction                                                      #
    # ==================================================================== #

    def _build_gui(self) -> None:
        # ── Title bar ──────────────────────────────────────────────────
        title_bar = tk.Frame(self.root, bg=self.C_CHROME, height=60)
        title_bar.pack(fill="x")
        title_bar.pack_propagate(False)

        if self._img_eee:
            tk.Label(title_bar, image=self._img_eee, bg=self.C_CHROME).pack(
                side="left", padx=(16, 8), pady=10)
        else:
            tk.Label(title_bar, text="EEE", font=("Montserrat", 10, "bold"),
                     bg=self.C_CHROME, fg="#000000").pack(side="left", padx=(16, 8), pady=10)

        tk.Label(title_bar, text="Course Audit Tool",
                 font=("Montserrat", 16, "bold"),
                 bg=self.C_CHROME, fg="#000000").pack(side="left", padx=8, pady=14)

        if self._img_uct:
            tk.Label(title_bar, image=self._img_uct, bg=self.C_CHROME).pack(
                side="right", padx=(8, 16), pady=10)
        else:
            tk.Label(title_bar, text="UCT", font=("Montserrat", 10, "bold"),
                     bg=self.C_CHROME, fg="#000000").pack(side="right", padx=(8, 16), pady=10)

        # ── Controls ───────────────────────────────────────────────────
        mf = tk.Frame(self.root, bg=self.C_CHROME)
        mf.pack(fill="x", padx=15, pady=10)

        def lbl(p, t):
            return tk.Label(p, text=t, bg=self.C_CHROME, fg=self.C_TEXT,
                            font=("Montserrat", 10))

        def btn(p, t, cmd, w=15):
            return tk.Button(p, text=t, command=cmd, width=w,
                             bg=self.C_INK, fg="#000000",
                             activebackground="#3A3536", activeforeground="#000000",
                             relief="flat", cursor="hand2",
                             font=("Montserrat", 10, "bold"), padx=6, pady=4)

        lbl(mf, "Recent Directories:").grid(row=0, column=0, sticky="w", pady=(4, 2))
        self.recent_combo = ttk.Combobox(mf, values=self.recent_directories,
                                         state="readonly", width=110)
        self.recent_combo.grid(row=1, column=0, sticky="ew", padx=(0, 8))
        self.recent_combo.bind("<<ComboboxSelected>>", self._select_recent)
        btn(mf, "Use Selected", self._use_recent).grid(row=1, column=1)

        lbl(mf, "Selected Course Root:").grid(row=2, column=0, sticky="w", pady=(12, 2))
        tk.Entry(mf, textvariable=self.selected_directory, width=110,
                 bg=self.C_CHROME, fg=self.C_TEXT, insertbackground=self.C_TEXT,
                 relief="flat", font=("Montserrat", 10)).grid(
                     row=3, column=0, sticky="ew", padx=(0, 8), ipady=4)
        btn(mf, "Browse...", self._browse).grid(row=3, column=1)

        sub = tk.Frame(mf, bg=self.C_CHROME)
        sub.grid(row=4, column=0, columnspan=2, sticky="w", pady=(12, 0))
        lbl(sub, "Folder structure profile:").grid(row=0, column=0, sticky="w", padx=(0, 40))
        lbl(sub, "Auditor:").grid(row=0, column=1, sticky="w", padx=(0, 40))
        lbl(sub, "Course TA:").grid(row=0, column=2, sticky="w")
        ttk.Combobox(sub, textvariable=self.profile_mode,
                     values=["Auto-detect", "Legacy", "Current", "New", "EEE4022"],
                     state="readonly", width=28).grid(
                         row=1, column=0, sticky="w", padx=(0, 40), pady=(4, 0))
        ttk.Combobox(sub, textvariable=self.selected_user,
                     values=AUDIT_USERS, state="readonly", width=28).grid(
                         row=1, column=1, sticky="w", padx=(0, 40), pady=(4, 0))
        ttk.Combobox(sub, textvariable=self.selected_ta,
                     values=TA_NAMES, width=28).grid(
                         row=1, column=2, sticky="w", pady=(4, 0))

        tk.Label(mf,
                 text=("Supports Legacy (2023/2024), Current (2025), and New (2026) "
                       "folder structures. GA courses are auto-detected. Folders "
                       "containing NONE are treated as intentionally empty."),
                 bg=self.C_CHROME, fg=self.C_TEXT,
                 font=("Montserrat", 9), anchor="w", justify="left").grid(
                     row=5, column=0, columnspan=2, sticky="w", pady=(8, 0))
        mf.columnconfigure(0, weight=1)

        # ── Button bar ─────────────────────────────────────────────────
        bf = tk.Frame(self.root, bg=self.C_CHROME)
        bf.pack(fill="x", padx=15, pady=(0, 8))
        btn(bf, "Run Audit and Create Outputs", self._run_audit, 28).pack(
            side="left", padx=(0, 8))
        btn(bf, "Clear Output", self._clear, 14).pack(side="left", padx=(0, 8))
        self.summary_label = tk.Label(bf, text="No audit run yet.",
                                      bg=self.C_CHROME, fg=self.C_TEXT,
                                      font=("Montserrat", 10), anchor="w")
        self.summary_label.pack(side="left", padx=12)

        # ── Notebook ───────────────────────────────────────────────────
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill="both", expand=True, padx=15, pady=(0, 12))

        self.output_tab   = tk.Frame(self.notebook, bg=self.C_CONTENT)
        self.issues_tab   = tk.Frame(self.notebook, bg=self.C_CONTENT)
        self.expected_tab = tk.Frame(self.notebook, bg=self.C_CONTENT)
        self.folder_tab   = tk.Frame(self.notebook, bg=self.C_CONTENT)
        self.file_tab     = tk.Frame(self.notebook, bg=self.C_CONTENT)
        self.tree_tab     = tk.Frame(self.notebook, bg=self.C_CONTENT)

        self.notebook.add(self.output_tab,   text="  Log Output  ")
        self.notebook.add(self.issues_tab,   text="  ⚠ Issues  ")
        self.notebook.add(self.expected_tab, text="  Expected Structure Check  ")
        self.notebook.add(self.folder_tab,   text="  Folder Details  ")
        self.notebook.add(self.file_tab,     text="  File Details  ")
        self.notebook.add(self.tree_tab,     text="  Tree Diagram  ")

        self._build_output_tab()
        self._build_issues_tab()
        self._build_expected_tab()
        self._build_folder_tab()
        self._build_file_tab()
        self._build_tree_tab()

    # ------------------------------------------------------------------ #
    #  Tab builders                                                         #
    # ------------------------------------------------------------------ #

    def _build_output_tab(self) -> None:
        self.output_text = scrolledtext.ScrolledText(
            self.output_tab, wrap=tk.WORD, width=140, height=32,
            bg=self.C_CONTENT, fg=self.C_TEXT,
            insertbackground=self.C_TEXT,
            selectbackground=self.C_INK, selectforeground=self.C_TEXT,
            font=("Montserrat", 10), relief="flat")
        self.output_text.pack(fill="both", expand=True, padx=2, pady=2)

    def _build_structure_tv(self, parent: tk.Frame) -> ttk.Treeview:
        """Build and return a Treeview with the standard expected-structure columns."""
        cols = ("relative_path", "level", "expected_name",
                "actual_name", "exists", "status", "details")
        tv = ttk.Treeview(parent, columns=cols, show="headings")
        headings = {
            "relative_path": "Parent Path", "level": "Level",
            "expected_name": "Expected Name", "actual_name": "Actual Name",
            "exists": "Exists", "status": "Status", "details": "Details",
        }
        widths = {
            "relative_path": 240, "level": 80, "expected_name": 250,
            "actual_name": 250, "exists": 70, "status": 170, "details": 340,
        }
        for col in cols:
            tv.heading(col, text=headings[col])
            tv.column(col, width=widths[col], anchor="w")
        ys = ttk.Scrollbar(parent, orient="vertical",   command=tv.yview)
        xs = ttk.Scrollbar(parent, orient="horizontal", command=tv.xview)
        tv.configure(yscrollcommand=ys.set, xscrollcommand=xs.set)
        tv.pack(side="left", fill="both", expand=True)
        ys.pack(side="right",  fill="y")
        xs.pack(side="bottom", fill="x")
        return tv

    def _build_issues_tab(self) -> None:
        self.issues_tv = self._build_structure_tv(self.issues_tab)

    def _build_expected_tab(self) -> None:
        self.expected_tv = self._build_structure_tv(self.expected_tab)

    def _build_folder_tab(self) -> None:
        cols = ("folder", "depth", "subfolder_count", "file_count",
                "total_size", "last_modified", "type_counts", "status")
        self.folder_tv = ttk.Treeview(self.folder_tab, columns=cols, show="headings")
        headings = {
            "folder": "Folder", "depth": "Depth", "subfolder_count": "Subfolders",
            "file_count": "Files", "total_size": "Size", "last_modified": "Latest Modified",
            "type_counts": "File Types", "status": "Status",
        }
        widths = {
            "folder": 320, "depth": 60, "subfolder_count": 80, "file_count": 60,
            "total_size": 100, "last_modified": 150, "type_counts": 340, "status": 180,
        }
        for col in cols:
            self.folder_tv.heading(col, text=headings[col])
            self.folder_tv.column(col, width=widths[col], anchor="w")
        ys = ttk.Scrollbar(self.folder_tab, orient="vertical",   command=self.folder_tv.yview)
        xs = ttk.Scrollbar(self.folder_tab, orient="horizontal", command=self.folder_tv.xview)
        self.folder_tv.configure(yscrollcommand=ys.set, xscrollcommand=xs.set)
        self.folder_tv.pack(side="left", fill="both", expand=True)
        ys.pack(side="right",  fill="y")
        xs.pack(side="bottom", fill="x")

    def _build_file_tab(self) -> None:
        cols = ("directory", "name", "type", "size", "modified")
        self.file_tv = ttk.Treeview(self.file_tab, columns=cols, show="headings")
        for col, title, width in [
            ("directory", "Directory", 320), ("name", "File Name", 300),
            ("type", "Type", 100), ("size", "Size", 100), ("modified", "Modified", 160),
        ]:
            self.file_tv.heading(col, text=title)
            self.file_tv.column(col, width=width, anchor="w")
        ys = ttk.Scrollbar(self.file_tab, orient="vertical",   command=self.file_tv.yview)
        xs = ttk.Scrollbar(self.file_tab, orient="horizontal", command=self.file_tv.xview)
        self.file_tv.configure(yscrollcommand=ys.set, xscrollcommand=xs.set)
        self.file_tv.pack(side="left", fill="both", expand=True)
        ys.pack(side="right",  fill="y")
        xs.pack(side="bottom", fill="x")

    def _build_tree_tab(self) -> None:
        self.tree_text = scrolledtext.ScrolledText(
            self.tree_tab, wrap=tk.NONE, width=140, height=32,
            bg=self.C_CONTENT, fg=self.C_TEXT,
            insertbackground=self.C_TEXT,
            selectbackground=self.C_CHROME, selectforeground=self.C_TEXT,
            font=("Montserrat", 10), relief="flat")
        self.tree_text.pack(fill="both", expand=True, padx=2, pady=2)

    # ==================================================================== #
    #  Logging                                                               #
    # ==================================================================== #

    def log(self, message: str) -> None:
        self.output_text.insert(tk.END, message + "\n")
        self.output_text.see(tk.END)
        self.root.update_idletasks()

    # ==================================================================== #
    #  File-system utilities                                                 #
    # ==================================================================== #

    def _folder_has_content(self, path: str) -> bool:
        """Return True if *path* (or any descendant) contains at least one real file."""
        try:
            for e in os.scandir(path):
                if e.name.lower() in IGNORED_FILES:
                    continue
                if e.is_file():
                    return True
                if e.is_dir() and self._folder_has_content(e.path):
                    return True
        except OSError:
            pass
        return False

    def _folder_has_direct_files(self, path: str) -> bool:
        """Return True if *path* contains at least one file directly (not in subfolders)."""
        try:
            for e in os.scandir(path):
                if e.is_file() and e.name.lower() not in IGNORED_FILES:
                    return True
        except OSError:
            pass
        return False

    def _latest_mtime(self, path: str) -> str:
        latest = None
        try:
            for e in os.scandir(path):
                if e.is_file() and e.name.lower() not in IGNORED_FILES:
                    try:
                        m = e.stat().st_mtime
                        if latest is None or m > latest:
                            latest = m
                    except OSError:
                        pass
        except OSError:
            pass
        return (datetime.fromtimestamp(latest).strftime("%Y-%m-%d %H:%M:%S")
                if latest else "N/A")

    @staticmethod
    def _fmt_size(size_bytes) -> str:
        if size_bytes is None:
            return "Unknown"
        for unit in ("B", "KB", "MB", "GB", "TB"):
            if size_bytes < 1024:
                return f"{size_bytes:.1f} {unit}"
            size_bytes /= 1024
        return f"{size_bytes:.1f} PB"

    @staticmethod
    def _rel(root: str, dirpath: str) -> str:
        return os.path.relpath(dirpath, root).replace("\\", "/")

    @staticmethod
    def _depth(rel: str) -> int:
        parts = rel.replace("\\", "/").split("/")
        return len([p for p in parts if p and p != "."])

    @staticmethod
    def _ext(filename: str) -> str:
        _, e = os.path.splitext(filename)
        return e.lower() if e else "[no ext]"

    # ==================================================================== #
    #  Submission validation                                                 #
    # ==================================================================== #

    SUBMISSION_EXTS: ClassVar[set[str]] = {".pdf", ".doc", ".docx"}

    def _is_submission_folder(self, name: str) -> bool:
        """Return True if *name* is a sample hand-ins, sample answers, or scripts folder."""
        nl = name.lower()
        is_sample  = "sample" in nl and ("hand" in nl or "answer" in nl)
        is_scripts = "script" in nl
        return is_sample or is_scripts

    def _is_under_submission(self, rel: str) -> bool:
        parts = rel.replace("\\", "/").split("/")
        return any(self._is_submission_folder(p) for p in parts)

    def _leaf_groups(self, path: str) -> list[tuple[str, list[str]]]:
        """Collect (folder_name, file_list) for each leaf directory under *path*."""
        groups: list[tuple[str, list[str]]] = []

        def _walk(p: str, name: str) -> None:
            try:
                entries = [e for e in os.scandir(p) if e.name.lower() not in IGNORED_FILES]
            except OSError:
                return
            files   = [e.name for e in entries if e.is_file()]
            subdirs = [e for e in entries if e.is_dir()]
            if files and not subdirs:
                groups.append((name, files))
            for sd in subdirs:
                _walk(sd.path, sd.name)

        _walk(path, os.path.basename(path))
        return groups

    def _check_submission(self, path: str) -> tuple[str, str]:
        """Validate a submission folder: each leaf group needs >=15 PDF/Word docs."""
        groups = self._leaf_groups(path)
        if not groups:
            return "EMPTY - REVIEW", "Subfolder exists but appears empty"

        issues: list[str] = []
        for name, files in groups:
            qualifying = [f for f in files
                          if os.path.splitext(f)[1].lower() in self.SUBMISSION_EXTS]
            count = len(qualifying)
            if count < 15:
                other = len(files) - count
                note  = f", {other} other type(s) ignored" if other else ""
                issues.append(f'"{name}": {count} PDF/Word doc(s) (expected >=15{note})')

        if not issues:
            total = sum(
                len([f for f in files if os.path.splitext(f)[1].lower() in self.SUBMISSION_EXTS])
                for _, files in groups
            )
            return "OK", (f"Submissions: {total} PDF/Word docs across "
                          f"{len(groups)} group{'s' if len(groups) != 1 else ''}")

        return "REVIEW - HAND-INS", (
            f"{len(issues)} group{'s' if len(issues) != 1 else ''} with issues: "
            + " | ".join(issues)
        )

    # ==================================================================== #
    #  Targeted content checks (course handout, DP list, etc.)              #
    # ==================================================================== #

    def _check_required_types(
        self, path: str, exts: set[str], friendly: str, label: str
    ) -> tuple[str, str]:
        all_files, matching = [], []
        for dp, _, fnames in os.walk(path):
            for fn in fnames:
                if fn.lower() in IGNORED_FILES:
                    continue
                all_files.append(fn)
                if os.path.splitext(fn)[1].lower() in exts:
                    matching.append(fn)
        if not all_files:
            return "EMPTY - REVIEW", f"{label}: folder is empty"
        if not matching:
            found = sorted({os.path.splitext(f)[1].lower() for f in all_files})
            return ("EMPTY - REVIEW",
                    f"{label}: no {friendly} found (found: {', '.join(found) or 'no ext'})")
        return "OK", f"{label}: {len(matching)} qualifying file{'s' if len(matching) != 1 else ''} found"

    def _is_course_handout(self, name: str) -> bool:
        return "course handout" in name.lower() or "course_handout" in name.lower()

    def _check_course_handout(self, path: str) -> tuple[str, str]:
        return self._check_required_types(
            path, {".pdf", ".doc", ".docx"}, "PDF or Word document", "Course handout")

    def _is_dp_list(self, name: str) -> bool:
        nl = name.lower()
        return "dp list" in nl or "dp_list" in nl

    def _check_dp_list(self, path: str) -> tuple[str, str]:
        return self._check_required_types(
            path, {".pdf", ".xls", ".xlsx", ".csv"}, "PDF or spreadsheet", "DP list")

    def _is_mark_sheets(self, name: str) -> bool:
        nl = name.lower()
        return "mark sheet" in nl or "marks sheet" in nl or "marksheet" in nl

    def _check_mark_sheets(self, path: str) -> tuple[str, str]:
        return self._check_required_types(
            path, {".pdf", ".xls", ".xlsx", ".csv"}, "PDF or spreadsheet", "Mark sheets")

    def _is_external_mod(self, name: str) -> bool:
        return "external moderator" in name.lower()

    def _check_external_mod(self, path: str) -> tuple[str, str]:
        return self._check_required_types(
            path, {".pdf", ".doc", ".docx"}, "PDF or Word document",
            "External moderator reports")

    # ==================================================================== #
    #  GA utilities                                                          #
    # ==================================================================== #

    def _is_ga_course(self, root_path: str) -> bool:
        folder = os.path.basename(root_path).strip()
        m = re.search(r'[A-Za-z]{2,}[0-9]{3,}[A-Za-z]?', folder)
        if not m:
            return False
        return m.group(0).lower() in GA_COURSES

    def _find_marksheets_folder(self, root_path: str) -> str | None:
        """Find the mark sheets folder under the main exam top-level folder.

        Strategy: scan ALL top-level folders whose name contains "exam",
        check each for a mark sheets subfolder, and return the first found.
        This handles all naming variants across profiles and years without
        assuming a specific folder name or number prefix.

        Also handles:
          - New profile 08_exams: drills into 01_main subgroup first
          - EEE4022: uses 3. Final Results as the reference folder

        Returns the absolute path to the mark sheets folder, or None.
        """
        try:
            top_dirs = [
                n for n in sorted(os.listdir(root_path), key=str.lower)
                if os.path.isdir(os.path.join(root_path, n))
            ]
        except OSError:
            return None

        # EEE4022: use 3. Final Results as reference
        for n in top_dirs:
            if nbk(n) == nbk("3. Final Results"):
                return os.path.join(root_path, n)

        # Search every top-level folder whose name contains "exam"
        # (covers 09. Exam, 12. Exams Main (Admin), 08_exams, etc.)
        for n in top_dirs:
            if "exam" not in n.lower():
                continue
            # Skip supplementary/SUPP folders — prefer main exam
            if any(kw in n.lower() for kw in ("supp", "supplementary", "de_")):
                continue

            exam_path = os.path.join(root_path, n)

            # New profile: 08_exams has sub-groups; drill into 01_main first
            try:
                children = [
                    c for c in sorted(os.listdir(exam_path), key=str.lower)
                    if os.path.isdir(os.path.join(exam_path, c))
                ]
            except OSError:
                continue

            # Check for a main sub-group (New profile 08_exams/01_main)
            main_group = next(
                (c for c in children
                 if nbk(c) in {"01_main", "main", "main exam", "main exam"}),
                None,
            )
            search_in = os.path.join(exam_path, main_group) if main_group else exam_path

            try:
                for c in os.listdir(search_in):
                    if (self._is_mark_sheets(c)
                            and os.path.isdir(os.path.join(search_in, c))):
                        return os.path.join(search_in, c)
            except OSError:
                continue

        # Fallback: try supplementary exam folder if nothing found above
        for n in top_dirs:
            if "exam" not in n.lower():
                continue
            exam_path = os.path.join(root_path, n)
            try:
                for c in os.listdir(exam_path):
                    if (self._is_mark_sheets(c)
                            and os.path.isdir(os.path.join(exam_path, c))):
                        return os.path.join(exam_path, c)
            except OSError:
                continue

        return None

    def _get_reference_student_count(self, root_path: str) -> int | None:
        """Return the reference student count for *root_path*.

        Reads from the mark sheets folder under the main exam folder.
        Looks for the first spreadsheet with an Emplid or Campus ID column
        and counts the data rows.

        Returns None if the folder or a suitable file cannot be found.
        """
        marksheets_folder = self._find_marksheets_folder(root_path)
        if marksheets_folder is None:
            return None
        return self._read_dp_count(marksheets_folder)

    def _read_dp_count(self, folder_path: str) -> int | None:
        """Find a spreadsheet in *folder_path* and return the student count.

        Since this is called with the mark sheets folder (or DP list folder)
        already resolved, we accept any xlsx/xls/csv file that contains an
        Emplid or Campus ID column.  We try each spreadsheet until one yields
        a valid count.
        """
        candidates: list[str] = []
        for dirpath, _, fnames in os.walk(folder_path):
            for fn in sorted(fnames):
                fl = fn.lower()
                if fl in IGNORED_FILES:
                    continue
                if os.path.splitext(fl)[1] in (".xlsx", ".xls", ".csv"):
                    candidates.append(os.path.join(dirpath, fn))

        for dp_file in candidates:
            count = self._count_from_file(dp_file)
            if count is not None:
                return count
        return None

    def _count_from_file(self, filepath: str) -> int | None:
        """Read *filepath* and return the count of student rows.

        Looks for a column whose lowercased header is in DP_LIST_STUDENT_COLUMNS
        (e.g. Emplid, Campus ID).  Returns None if the file cannot be read or
        no matching column is found.
        """
        ext = os.path.splitext(filepath)[1].lower()
        try:
            if ext == ".csv":
                with open(filepath, "r", encoding="utf-8", errors="replace") as f:
                    reader = csv.reader(f)
                    headers = [h.strip().lower() for h in next(reader, [])]
                    col = next((i for i, h in enumerate(headers)
                                if h in DP_LIST_STUDENT_COLUMNS), None)
                    if col is None:
                        return None
                    return sum(1 for row in reader
                               if len(row) > col and row[col].strip())
            else:
                from openpyxl import load_workbook
                wb   = load_workbook(filepath, read_only=True, data_only=True)
                ws   = wb.active
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    return None
                headers = [str(h).strip().lower() if h is not None else ""
                           for h in rows[0]]
                col = next((i for i, h in enumerate(headers)
                            if h in DP_LIST_STUDENT_COLUMNS), None)
                if col is None:
                    return None
                return sum(1 for row in rows[1:]
                           if len(row) > col
                           and row[col] is not None
                           and str(row[col]).strip())
        except Exception:
            return None

    def _check_ga_assessment_forms(
        self, path: str, ga_folder_path: str
    ) -> tuple[str, str]:
        """Validate the GA assessment forms subfolder against the DP list count."""
        if not self._folder_has_content(path):
            return "EMPTY - REVIEW", "GA assessment forms folder is empty"

        # Check if it's a single combined PDF
        all_files: list[str] = []
        for dp, _, fnames in os.walk(path):
            for fn in fnames:
                if fn.lower() not in IGNORED_FILES:
                    all_files.append(fn)

        qualifying = [f for f in all_files
                      if os.path.splitext(f)[1].lower() in self.SUBMISSION_EXTS]

        if len(qualifying) == 1:
            return ("OK",
                    "GA assessment forms: single combined PDF found — "
                    "verify student count manually")

        # Use the reference student count from the mark sheets folder
        student_count = getattr(self, "_ref_student_count", None)

        if student_count is None:
            return ("OK",
                    f"GA assessment forms: {len(qualifying)} PDF/Word file(s) found "
                    "(reference student count not available — verify manually)")

        if len(qualifying) >= student_count:
            return ("OK",
                    f"GA assessment forms: {len(qualifying)} files, "
                    f"{student_count} students — complete")

        missing = student_count - len(qualifying)
        return ("REVIEW - GA INCOMPLETE",
                f"GA assessment forms: {len(qualifying)} files found, "
                f"{student_count} students on DP list — {missing} submission(s) missing")

    # ==================================================================== #
    #  Result row factory                                                    #
    # ==================================================================== #

    def _row(
        self,
        relative_path: str,
        level: str,
        expected_name: str,
        actual_name: str,
        exists: str,
        status: str,
        details: str,
    ) -> dict:
        return {
            "relative_path": relative_path,
            "level":         level,
            "expected_name": expected_name,
            "actual_name":   actual_name,
            "exists":        exists,
            "status":        status,
            "details":       details,
        }

    # ==================================================================== #
    #  Profile detection                                                     #
    # ==================================================================== #

    def _detect_profile(self, root_path: str) -> str:
        """Determine the folder-structure profile for *root_path*.

        Detection order:
          1. Manual override from dropdown
          2. Year from folder name (≤2024 → Legacy, 2025 → Current, ≥2026 → New)
          3. Snake_case folder names → New
          4. Disk marker scoring
          5. Default → Current
        """
        mode = self.profile_mode.get()
        if mode in STRUCTURE_PROFILES:
            return mode

        folder_name = os.path.basename(root_path)

        # Year extraction
        year_m = re.search(r'\b(20\d{2})\b', folder_name)
        if year_m:
            year = int(year_m.group(1))
            if year >= 2026:
                return "New"
            if year == 2025:
                return "Current"
            if year <= 2024:
                return "Legacy"

        # EEE4022 project-based course — detected by course code alone
        if re.search(r'EEE4022', folder_name, re.IGNORECASE):
            return "EEE4022"

        # Snake_case detection (New profile)
        try:
            top_dirs = [n for n in os.listdir(root_path)
                        if os.path.isdir(os.path.join(root_path, n))]
        except OSError:
            return "Current"

        if any("_" in n and "." not in n for n in top_dirs):
            return "New"

        # Marker scoring
        names = {nbk(n) for n in top_dirs}
        legacy_markers  = {"exam", "supplementary exam"}
        current_markers = {"exams main", "exams supp"}

        legacy_score  = sum(1 for m in legacy_markers  if m in names)
        current_score = sum(1 for m in current_markers if m in names)

        if legacy_score > current_score:
            return "Legacy"
        if current_score > 0:
            return "Current"
        return "Current"

    # ==================================================================== #
    #  Structure evaluation helpers                                          #
    # ==================================================================== #

    def _check_duplicates(
        self, entries: list[str], rel: str, level: str, results: list[dict]
    ) -> None:
        """Flag folders at the same level whose normalised key collides."""
        counts = Counter(nbk(n) for n in entries)
        for key, count in counts.items():
            if count > 1:
                dupes = [n for n in entries if nbk(n) == key]
                results.append(self._row(
                    rel, level, "", ", ".join(dupes), "Yes", "DUPLICATE",
                    f"Multiple folders resolve to the same name: {', '.join(dupes)}",
                ))

    def _evaluate_child(
        self,
        child_name: str,
        actual_child: str,
        child_path: str,
        parent_display: str,
        results: list[dict],
    ) -> None:
        """Evaluate a single expected subfolder and append a result row."""
        has_content = self._folder_has_content(child_path)
        is_none     = _has_none(actual_child)
        has_marker  = _has_admin_marker(actual_child)

        if is_none:
            status = "NONE - ACCEPTED" if not has_content else "POPULATED DESPITE NONE"
            detail = "Subfolder is marked NONE" + (
                " but contains files or subfolders" if has_content else "")
        elif has_marker:
            status = "ADMIN FLAG"
            detail = ("Administrator status marker in subfolder name — "
                      + ("has content" if has_content else "empty")
                      + ", please review and rename")
        elif not has_content:
            status, detail = "EMPTY - REVIEW", "Subfolder exists but appears empty"
        elif self._is_submission_folder(child_name):
            status, detail = self._check_submission(child_path)
        elif self._is_course_handout(child_name):
            status, detail = self._check_course_handout(child_path)
        elif self._is_dp_list(child_name):
            status, detail = self._check_dp_list(child_path)
        elif self._is_mark_sheets(child_name):
            status, detail = self._check_mark_sheets(child_path)
        elif self._is_external_mod(child_name):
            status, detail = self._check_external_mod(child_path)
        else:
            status, detail = "OK", "Subfolder found"

        results.append(self._row(
            parent_display, "Subfolder", child_name, actual_child,
            "Yes", status, detail,
        ))

    def _evaluate_children(
        self,
        top_path: str,
        parent_display: str,
        expected_children: list[str],
        actual_children: list[str],
        results: list[dict],
    ) -> None:
        """Compare one level of subfolders against the template expected list."""
        actual_map  = {nbk(n): n for n in actual_children}
        expected_keys = set()

        for exp in expected_children:
            key = nbk(exp)
            expected_keys.add(key)
            actual = actual_map.get(key)

            if actual is None:
                results.append(self._row(
                    parent_display, "Subfolder", exp, "",
                    "No", "MISSING", "Expected subfolder is missing",
                ))
                continue

            self._evaluate_child(
                exp, actual, os.path.join(top_path, actual),
                parent_display, results,
            )

        # Flag unexpected subfolders
        for actual in actual_children:
            if nbk(actual) not in expected_keys:
                base = nbk(actual)
                detail = (
                    "Subfolder name matches a known template folder but may be "
                    "renumbered or renamed — check against the expected structure"
                    if base in _ALL_TEMPLATE_BASE_NAMES
                    else "Subfolder exists but is not in the template"
                )
                results.append(self._row(
                    parent_display, "Subfolder", "", actual,
                    "Yes", "UNEXPECTED", detail,
                ))

    def _evaluate_two_level_exam(
        self,
        exam_path: str,
        exam_display: str,
        exam_groups: list[str],
        subfolder_template: list[str],
        results: list[dict],
    ) -> None:
        """Validate a two-level exam folder (New profile 08_exams)."""
        try:
            actual_children = [
                n for n in sorted(os.listdir(exam_path), key=str.lower)
                if os.path.isdir(os.path.join(exam_path, n))
            ]
        except OSError:
            actual_children = []

        actual_map    = {nbk(n): n for n in actual_children}
        expected_keys = set()

        for exp_group in exam_groups:
            key = nbk(exp_group)
            expected_keys.add(key)
            actual_group = actual_map.get(key)

            if actual_group is None:
                results.append(self._row(
                    exam_display, "Subfolder", exp_group, "",
                    "No", "MISSING", "Expected exam group is missing",
                ))
                continue

            group_path  = os.path.join(exam_path, actual_group)
            has_content = self._folder_has_content(group_path)
            is_none     = _has_none(actual_group)

            if is_none:
                status = "NONE - ACCEPTED" if not has_content else "POPULATED DESPITE NONE"
                detail = "Exam group is marked NONE" + (
                    " but contains files" if has_content else "")
                results.append(self._row(
                    exam_display, "Subfolder", exp_group, actual_group,
                    "Yes", status, detail))
                continue

            results.append(self._row(
                exam_display, "Subfolder", exp_group, actual_group,
                "Yes",
                "OK" if has_content else "EMPTY - REVIEW",
                "Exam group found" if has_content else "Exam group exists but is empty",
            ))

            group_display = f"{exam_display}/{_strip_none(actual_group)}"
            try:
                grandchildren = [
                    n for n in sorted(os.listdir(group_path), key=str.lower)
                    if os.path.isdir(os.path.join(group_path, n))
                ]
            except OSError:
                grandchildren = []

            self._evaluate_children(
                group_path, group_display, subfolder_template, grandchildren, results)
            self._check_duplicates(grandchildren, group_display, "Subfolder", results)

        for actual in actual_children:
            if nbk(actual) not in expected_keys:
                results.append(self._row(
                    exam_display, "Subfolder", "", actual,
                    "Yes", "UNEXPECTED", "Subfolder is not in the exam template",
                ))
        self._check_duplicates(actual_children, exam_display, "Subfolder", results)

    def _evaluate_ga_moderation(
        self,
        ga_path: str,
        ga_name: str,
        results: list[dict],
    ) -> None:
        """Validate the GA moderation folder and its subfolders."""
        is_new = "_" in ga_name
        expected_subs = NEW_GA_MOD_SUBFOLDERS if is_new else LEGACY_GA_MOD_SUBFOLDERS
        display = _strip_none(ga_name)

        try:
            actual_children = [
                n for n in sorted(os.listdir(ga_path), key=str.lower)
                if os.path.isdir(os.path.join(ga_path, n))
            ]
        except OSError:
            actual_children = []

        actual_map    = {nbk(n): n for n in actual_children}
        expected_keys = set()

        for exp in expected_subs:
            key = nbk(exp)
            expected_keys.add(key)
            actual = actual_map.get(key)

            if actual is None:
                results.append(self._row(
                    display, "Subfolder", exp, "",
                    "No", "MISSING", "Expected GA moderation subfolder is missing",
                ))
                continue

            sub_path    = os.path.join(ga_path, actual)
            has_content = self._folder_has_content(sub_path)
            is_none     = _has_none(actual)

            if is_none:
                status = "NONE - ACCEPTED" if not has_content else "POPULATED DESPITE NONE"
                detail = "Subfolder is marked NONE" + (
                    " but contains files" if has_content else "")
                results.append(self._row(
                    display, "Subfolder", exp, actual, "Yes", status, detail))
                continue

            # GA assessment forms — special student count check
            is_assessment = (
                "assessment form" in exp.lower()
                or "assessment_form" in exp.lower()
            )
            if is_assessment:
                status, detail = self._check_ga_assessment_forms(sub_path, ga_path)
            elif not has_content:
                status, detail = "EMPTY - REVIEW", "GA subfolder exists but appears empty"
            else:
                status, detail = "OK", "GA moderation subfolder found"

            results.append(self._row(
                display, "Subfolder", exp, actual, "Yes", status, detail))

        for actual in actual_children:
            if nbk(actual) not in expected_keys:
                results.append(self._row(
                    display, "Subfolder", "", actual,
                    "Yes", "UNEXPECTED", "Not in the GA moderation template",
                ))

        self._check_duplicates(actual_children, display, "Subfolder", results)

    # ==================================================================== #
    #  Main structure evaluation                                             #
    # ==================================================================== #

    # ==================================================================== #
    #  EEE4022 project validation                                           #
    # ==================================================================== #

    def _check_student_project_folder(
        self, student_path: str, student_name: str, parent_display: str,
        results: list[dict],
    ) -> None:
        """Validate a single student project subfolder.

        Rules:
          - Must contain a Final Report PDF (filename contains "final report",
            case-insensitive)
          - Must have at least one marker report PDF
        """
        try:
            files = [
                f for f in os.listdir(student_path)
                if os.path.isfile(os.path.join(student_path, f))
                and f.lower() not in IGNORED_FILES
            ]
        except OSError:
            files = []

        if not files:
            results.append(self._row(
                parent_display, "Subfolder", student_name, student_name,
                "Yes", "EMPTY - REVIEW", "Student project folder is empty",
            ))
            return

        has_final_report = any(
            "final report" in f.lower() and f.lower().endswith(".pdf")
            for f in files
        )
        has_marker = any(
            "marker" in f.lower() and f.lower().endswith(".pdf")
            for f in files
        )

        if not has_final_report and not has_marker:
            results.append(self._row(
                parent_display, "Subfolder", student_name, student_name,
                "Yes", "REVIEW - HAND-INS",
                f"{student_name}: no Final Report PDF and no marker report PDF found",
            ))
        elif not has_final_report:
            results.append(self._row(
                parent_display, "Subfolder", student_name, student_name,
                "Yes", "REVIEW - HAND-INS",
                f"{student_name}: Final Report PDF is missing",
            ))
        else:
            results.append(self._row(
                parent_display, "Subfolder", student_name, student_name,
                "Yes", "OK",
                f"{student_name}: Final Report and marker report(s) found "
                f"({len(files)} file(s) total)",
            ))

    def _check_moderator_folder(
        self, mod_path: str, mod_name: str, parent_display: str,
        results: list[dict],
    ) -> None:
        """Validate a single moderator subfolder.

        Rules:
          - Must contain a Part A PDF
          - Must contain at least one Part B PDF (or Word doc)
        """
        try:
            files = [
                f for f in os.listdir(mod_path)
                if os.path.isfile(os.path.join(mod_path, f))
                and f.lower() not in IGNORED_FILES
            ]
        except OSError:
            files = []

        if not files:
            results.append(self._row(
                parent_display, "Subfolder", mod_name, mod_name,
                "Yes", "EMPTY - REVIEW", "Moderator folder is empty",
            ))
            return

        pdf_doc_files = [
            f for f in files
            if os.path.splitext(f)[1].lower() in {".pdf", ".doc", ".docx"}
        ]
        has_part_a = any("part a" in f.lower() for f in pdf_doc_files)
        part_b_files = [f for f in pdf_doc_files if "part b" in f.lower()]

        if not has_part_a and not part_b_files:
            status = "REVIEW - HAND-INS"
            detail = f"{mod_name}: no Part A and no Part B files found"
        elif not has_part_a:
            status = "REVIEW - HAND-INS"
            detail = f"{mod_name}: Part A is missing ({len(part_b_files)} Part B file(s) found)"
        elif not part_b_files:
            status = "REVIEW - HAND-INS"
            detail = f"{mod_name}: Part A found but no Part B files found"
        else:
            status = "OK"
            detail = (f"{mod_name}: Part A and {len(part_b_files)} Part B file(s) found")

        results.append(self._row(
            parent_display, "Subfolder", mod_name, mod_name,
            "Yes", status, detail,
        ))

    def _evaluate_eee4022(self, root_path: str) -> list[dict]:
        """Full structure evaluation for the EEE4022 project-based profile.

        Validates:
          1. Top-level folders against EEE4022_STRUCTURE template
          2. Each student subfolder in '4. Student Projects' for Final Report
          3. Each moderator subfolder in '5. Moderators' for Part A + Part B
          4. GA moderation folder (compulsory — EEE4022 is always a GA course)
          5. Unexpected top-level folders
        """
        results: list[dict] = []

        try:
            top_entries = [
                n for n in sorted(os.listdir(root_path), key=str.lower)
                if os.path.isdir(os.path.join(root_path, n))
            ]
        except OSError:
            return results

        top_map  = {nbk(n): n for n in top_entries}
        top_keys = set()

        # GA moderation is always compulsory for EEE4022
        ga_folder = next(
            (n for n in top_entries if nbk(n) in _GA_MOD_KEYS), None)
        if ga_folder is None:
            results.append(self._row(
                "", "Top level", "00_ga_moderation", "",
                "No", "MISSING",
                "GA moderation folder is compulsory for EEE4022 but is missing",
            ))
        else:
            ga_path     = os.path.join(root_path, ga_folder)
            has_content = self._folder_has_content(ga_path)
            results.append(self._row(
                "", "Top level", "00_ga_moderation", ga_folder,
                "Yes",
                "OK" if has_content else "EMPTY - REVIEW",
                "GA moderation folder found"
                if has_content else "GA moderation folder exists but is empty",
            ))
            self._evaluate_ga_moderation(ga_path, ga_folder, results)
        for n in top_entries:
            if nbk(n) in _GA_MOD_KEYS:
                top_keys.add(nbk(n))

        # Standard EEE4022 top-level folders
        for exp_top in EEE4022_STRUCTURE:
            key = nbk(exp_top)
            top_keys.add(key)
            actual = top_map.get(key)

            if actual is None:
                results.append(self._row(
                    "", "Top level", exp_top, "",
                    "No", "MISSING", "Expected EEE4022 folder is missing",
                ))
                continue

            top_path    = os.path.join(root_path, actual)
            has_content = self._folder_has_content(top_path)
            is_none     = _has_none(actual)
            has_marker  = _has_admin_marker(actual)

            if is_none:
                status = "NONE - ACCEPTED" if not has_content else "POPULATED DESPITE NONE"
                detail = "Folder is marked NONE" + (
                    " but contains files or subfolders" if has_content else "")
            elif has_marker:
                status, detail = "ADMIN FLAG", "Administrator status marker in folder name"
            elif not has_content:
                status, detail = "EMPTY - REVIEW", "Folder exists but contains no files"
            else:
                status, detail = "OK", "Folder found"

            results.append(self._row(
                "", "Top level", exp_top, actual, "Yes", status, detail))

            if is_none:
                continue

            # Per-student validation
            if nbk(actual) == nbk("4. Student Projects"):
                try:
                    students = [
                        n for n in sorted(os.listdir(top_path), key=str.lower)
                        if os.path.isdir(os.path.join(top_path, n))
                    ]
                except OSError:
                    students = []
                display = _strip_none(actual)
                for student in students:
                    self._check_student_project_folder(
                        os.path.join(top_path, student),
                        student, display, results,
                    )
                self._check_duplicates(students, display, "Subfolder", results)
                continue

            # Per-moderator validation
            if nbk(actual) == nbk("5. Moderators"):
                try:
                    mods = [
                        n for n in sorted(os.listdir(top_path), key=str.lower)
                        if os.path.isdir(os.path.join(top_path, n))
                    ]
                except OSError:
                    mods = []
                display = _strip_none(actual)
                for mod in mods:
                    self._check_moderator_folder(
                        os.path.join(top_path, mod),
                        mod, display, results,
                    )
                self._check_duplicates(mods, display, "Subfolder", results)
                continue

        # Unexpected top-level folders
        for actual in top_entries:
            if nbk(actual) not in top_keys:
                results.append(self._row(
                    "", "Top level", "", actual,
                    "Yes", "UNEXPECTED",
                    "Folder exists but is not in the EEE4022 template",
                ))

        self._check_duplicates(top_entries, "", "Top level", results)
        return results

    def evaluate_structure(self, root_path: str, profile_name: str) -> list[dict]:
        """Compare *root_path* against *profile_name* template. Return result rows."""
        expected = STRUCTURE_PROFILES[profile_name]
        results:  list[dict] = []

        # EEE4022 has its own dedicated evaluator
        if profile_name == "EEE4022":
            return self._evaluate_eee4022(root_path)

        try:
            top_entries = [
                n for n in sorted(os.listdir(root_path), key=str.lower)
                if os.path.isdir(os.path.join(root_path, n))
            ]
        except OSError:
            return results

        top_map    = {nbk(n): n for n in top_entries}
        top_keys   = set()

        # ── GA moderation (compulsory for GA courses) ──────────────────
        if self._is_ga_course(root_path):
            ga_folder = next(
                (n for n in top_entries if nbk(n) in _GA_MOD_KEYS), None)
            if ga_folder is None:
                results.append(self._row(
                    "", "Top level", "00_ga_moderation", "",
                    "No", "MISSING",
                    "GA moderation folder is compulsory for this course but is missing",
                ))
            else:
                ga_path     = os.path.join(root_path, ga_folder)
                has_content = self._folder_has_content(ga_path)
                results.append(self._row(
                    "", "Top level", "00_ga_moderation", ga_folder,
                    "Yes",
                    "OK" if has_content else "EMPTY - REVIEW",
                    "GA moderation folder found"
                    if has_content else "GA moderation folder exists but is empty",
                ))
                self._evaluate_ga_moderation(ga_path, ga_folder, results)
            # Exclude from unexpected check
            for n in top_entries:
                if nbk(n) in _GA_MOD_KEYS:
                    top_keys.add(nbk(n))

        # ── Standard profile evaluation ────────────────────────────────
        for exp_top, exp_children in expected.items():
            key = nbk(exp_top)
            top_keys.add(key)
            actual_top = top_map.get(key)

            if actual_top is None:
                results.append(self._row(
                    "", "Top level", exp_top, "",
                    "No", "MISSING", "Expected top-level folder is missing",
                ))
                continue

            top_path    = os.path.join(root_path, actual_top)
            has_content = self._folder_has_content(top_path)
            is_none     = _has_none(actual_top)
            has_marker  = _has_admin_marker(actual_top)

            if is_none:
                status = "NONE - ACCEPTED" if not has_content else "POPULATED DESPITE NONE"
                detail = "Folder is marked NONE" + (
                    " but contains files or subfolders" if has_content else "")
            elif has_marker:
                status = "ADMIN FLAG"
                detail = ("Administrator status marker in folder name — "
                          + ("has content" if has_content else "empty")
                          + ", please review and rename")
            elif has_content:
                status, detail = "OK", "Top-level folder found"
            else:
                status, detail = "EMPTY - REVIEW", "Top-level folder exists but contains no files"

            results.append(self._row(
                "", "Top level", exp_top, actual_top, "Yes", status, detail))

            if is_none:
                # Surface children of NONE-marked top folder
                try:
                    for child in sorted(os.listdir(top_path), key=str.lower):
                        if os.path.isdir(os.path.join(top_path, child)):
                            ch = self._folder_has_content(os.path.join(top_path, child))
                            results.append(self._row(
                                _strip_none(actual_top), "Subfolder", "", child, "Yes",
                                "NONE - ACCEPTED" if not ch else "POPULATED DESPITE NONE",
                                "Child of a NONE-marked folder",
                            ))
                except OSError:
                    pass
                continue

            # New profile: 08_exams uses two-level structure
            if profile_name == "New" and nbk(exp_top) == nbk("08_exams"):
                self._evaluate_two_level_exam(
                    top_path, _strip_none(actual_top),
                    exp_children, NEW_EXAM_SUBFOLDERS, results,
                )
                continue

            # Flat-file tolerance
            if exp_children and self._folder_has_direct_files(top_path):
                results[-1]["status"]  = "OK"
                results[-1]["details"] = "Top-level folder contains files directly (no subfolders)"
                continue

            try:
                actual_children = [
                    n for n in sorted(os.listdir(top_path), key=str.lower)
                    if os.path.isdir(os.path.join(top_path, n))
                ]
            except OSError:
                actual_children = []

            self._evaluate_children(
                top_path, _strip_none(actual_top),
                exp_children, actual_children, results,
            )
            self._check_duplicates(
                actual_children, _strip_none(actual_top), "Subfolder", results)

        # ── Unexpected top-level folders ───────────────────────────────
        for actual in top_entries:
            if nbk(actual) not in top_keys:
                base = nbk(actual)
                detail = (
                    "Folder name matches a known template folder but may be "
                    "renumbered or renamed — check against the expected structure"
                    if base in _ALL_TEMPLATE_BASE_NAMES
                    else "Folder exists but is not in the template"
                )
                results.append(self._row(
                    "", "Top level", "", actual, "Yes", "UNEXPECTED", detail))

        self._check_duplicates(top_entries, "", "Top level", results)
        return results

    # ==================================================================== #
    #  Folder status for Folder Details tab                                  #
    # ==================================================================== #

    def _folder_status(
        self, rel: str, expected_results: list[dict], abs_path: str = ""
    ) -> str:
        """Derive the display status for a folder row in Folder Details."""
        folder_name = os.path.basename(rel)
        parent      = "/".join(rel.replace("\\", "/").split("/")[:-1])

        relevant = [
            r for r in expected_results
            if r["actual_name"] == folder_name and (
                (r["relative_path"] == "" and parent == "")
                or (r["relative_path"] != "" and r["relative_path"] == parent)
                or (r["relative_path"] != "" and
                    rel.startswith(r["relative_path"] + "/"))
            )
        ]
        statuses = {r["status"] for r in relevant}

        for p in [
            "DUPLICATE", "ADMIN FLAG", "POPULATED DESPITE NONE",
            "REVIEW - HAND-INS", "REVIEW - GA INCOMPLETE",
            "MISSING", "EMPTY - REVIEW", "NONE - ACCEPTED", "UNEXPECTED",
        ]:
            if p in statuses:
                return "MISSING CHILDREN" if p == "MISSING" else p
        if statuses == {"OK"}:
            return "OK"

        # Deep subfolder not in expected results — evaluate directly
        if abs_path:
            if _has_none(folder_name):
                return ("NONE - ACCEPTED" if not self._folder_has_content(abs_path)
                        else "POPULATED DESPITE NONE")
            if not self._folder_has_content(abs_path):
                return "EMPTY - REVIEW"
            if self._is_under_submission(rel):
                status, _ = self._check_submission(abs_path)
                return status
            return "OK"
        return "OK"

    # ==================================================================== #
    #  ASCII tree                                                            #
    # ==================================================================== #

    def _ascii_tree(self, root_path: str) -> str:
        lines: list[str] = []

        def _walk(path: str, prefix: str) -> None:
            try:
                entries = sorted(os.scandir(path),
                                 key=lambda e: (e.is_file(), e.name.lower()))
            except OSError:
                return
            entries = [e for e in entries if e.name.lower() not in IGNORED_FILES]
            for i, e in enumerate(entries):
                is_last   = i == len(entries) - 1
                connector = "└── " if is_last else "├── "
                lines.append(prefix + connector + e.name)
                if e.is_dir():
                    _walk(e.path, prefix + ("    " if is_last else "│   "))

        lines.append(os.path.basename(root_path) or root_path)
        _walk(root_path, "")
        return "\n".join(lines)

    # ==================================================================== #
    #  Core analysis                                                         #
    # ==================================================================== #

    def analyse(self, root_path: str) -> dict:
        total_files = total_folders = total_size = 0
        folder_data:         list[dict] = []
        file_details:        list[dict] = []
        unsigned_issues:     list[dict] = []
        overall_file_types             = Counter()

        profile          = self._detect_profile(root_path)
        # Compute reference student count once from the mark sheets folder
        # and cache it for use throughout this audit run.
        _ms_folder = self._find_marksheets_folder(root_path)
        self._ref_student_count = self._get_reference_student_count(root_path)
        if self._ref_student_count is not None:
            self.log(f"Reference student count: {self._ref_student_count} "
                     f"(from mark sheets folder)")
        elif _ms_folder:
            self.log("Reference student count: not found — mark sheets folder "
                     f"found at {os.path.basename(_ms_folder)!r} but no spreadsheet "
                     "with Emplid/Campus ID column — add a UCT class list there")
        else:
            self.log("Reference student count: not found — no mark sheets folder located")
        expected_results = self.evaluate_structure(root_path, profile)

        for dirpath, dirnames, filenames in os.walk(root_path):
            dirnames.sort()
            filenames.sort()
            dirnames[:]   = [d for d in dirnames   if d.lower() not in IGNORED_FILES]
            filenames_flt = [f for f in filenames  if f.lower() not in IGNORED_FILES]

            if dirpath == root_path:
                continue

            total_folders += 1
            rel            = self._rel(root_path, dirpath)
            folder_size    = 0
            type_counter   = Counter()

            for fn in filenames_flt:
                total_files += 1
                ext       = self._ext(fn)
                full_path = os.path.join(dirpath, fn)
                try:
                    size = os.path.getsize(full_path)
                except OSError:
                    size = None
                try:
                    mtime = datetime.fromtimestamp(
                        os.path.getmtime(full_path)).strftime("%Y-%m-%d %H:%M:%S")
                except OSError:
                    mtime = "Unavailable"

                if size is not None:
                    total_size   += size
                    folder_size  += size

                type_counter[ext]        += 1
                overall_file_types[ext]  += 1

                file_details.append({
                    "directory": rel, "name": fn,
                    "extension": ext, "size": size, "modified": mtime,
                })

                stem = os.path.splitext(fn)[0]
                if _UNSIGNED_PATTERN.search(stem):
                    unsigned_issues.append(self._row(
                        rel, "File", "", fn, "Yes", "UNSIGNED FILE",
                        "File name indicates it has not been signed — please review",
                    ))

            folder_data.append({
                "directory":              rel,
                "depth":                  self._depth(rel),
                "subfolder_count":        len(dirnames),
                "file_count":             len(filenames_flt),
                "folder_file_total_size": folder_size,
                "latest_modified":        self._latest_mtime(dirpath),
                "file_type_counts":       dict(sorted(type_counter.items())),
                "status":                 self._folder_status(rel, expected_results, dirpath),
            })

        issue_rows = (
            [r for r in expected_results if r["status"] in self._ISSUE_STATUSES]
            + unsigned_issues
        )

        return {
            "profile":             profile,
            "is_ga":               self._is_ga_course(root_path),
            "total_files":         total_files,
            "total_folders":       total_folders,
            "total_size_bytes":    total_size,
            "folder_data":         folder_data,
            "file_details":        file_details,
            "overall_file_types":  dict(sorted(overall_file_types.items())),
            "ascii_tree":          self._ascii_tree(root_path),
            "expected_results":    expected_results,
            "issues":              issue_rows,
        }

    # ==================================================================== #
    #  Populate GUI tables                                                   #
    # ==================================================================== #

    def _populate_structure_tv(self, tv: ttk.Treeview, rows: list[dict]) -> None:
        for item in tv.get_children():
            tv.delete(item)
        for r in rows:
            tv.insert("", tk.END, values=(
                r["relative_path"], r["level"], r["expected_name"],
                r["actual_name"],   r["exists"], r["status"], r["details"],
            ))

    def _populate_all(self, data: dict) -> None:
        self._populate_structure_tv(self.issues_tv,   data["issues"])
        self._populate_structure_tv(self.expected_tv, data["expected_results"])

        for item in self.folder_tv.get_children():
            self.folder_tv.delete(item)
        for f in data["folder_data"]:
            tc = (", ".join(f"{k}: {v}" for k, v in f["file_type_counts"].items())
                  if f["file_type_counts"] else "None")
            self.folder_tv.insert("", tk.END, values=(
                f["directory"], f["depth"], f["subfolder_count"], f["file_count"],
                self._fmt_size(f["folder_file_total_size"]),
                f["latest_modified"], tc, f["status"],
            ))

        for item in self.file_tv.get_children():
            self.file_tv.delete(item)
        for fi in data["file_details"]:
            self.file_tv.insert("", tk.END, values=(
                fi["directory"], fi["name"], fi["extension"],
                self._fmt_size(fi["size"]), fi["modified"],
            ))

        self.tree_text.delete("1.0", tk.END)
        self.tree_text.insert(tk.END, data["ascii_tree"])
        self.notebook.select(self.issues_tab)

    # ==================================================================== #
    #  File output                                                           #
    # ==================================================================== #

    def _output_stem(self, base_path: str) -> str:
        date     = datetime.now().strftime("%Y%m%d")
        folder   = os.path.basename(base_path).strip()
        m        = re.search(r'[A-Za-z]{2,}[0-9]{3,}[A-Za-z]?', folder)
        code     = m.group(0) if m else re.sub(r'[^\w\-.]', '_', folder).strip('_')
        return f"{date}_{code}".lower()

    def _write_log(self, path: str, data: dict, root: str) -> None:
        with open(path, "w", encoding="utf-8") as f:
            f.write("COURSE FOLDER AUDIT LOG\n" + "=" * 100 + "\n\n")
            f.write("SCAN DETAILS\n" + "-" * 100 + "\n")
            f.write(f"Root directory:   {root}\n")
            f.write(f"Auditor:          {self.selected_user.get()}\n")
            if self.selected_ta.get().strip():
                f.write(f"Course TA:        {self.selected_ta.get().strip()}\n")
            f.write(f"Detected profile: {data['profile']}"
                    f"{'  [GA course]' if data['is_ga'] else ''}\n")
            f.write(f"Scan date:        {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            f.write("SUMMARY\n" + "-" * 100 + "\n")
            f.write(f"Total folders: {data['total_folders']}\n")
            f.write(f"Total files:   {data['total_files']}\n")
            f.write(f"Total size:    {self._fmt_size(data['total_size_bytes'])}\n")
            f.write(f"Issues found:  {len(data['issues'])}\n\n")
            f.write("TREE STRUCTURE\n" + "-" * 100 + "\n")
            f.write(data["ascii_tree"] + "\n\n")
            f.write("EXPECTED STRUCTURE CHECK\n" + "-" * 100 + "\n")
            for r in data["expected_results"]:
                f.write(
                    f"Parent: {r['relative_path'] or '[root]'} | "
                    f"Level: {r['level']} | "
                    f"Expected: {r['expected_name'] or '-'} | "
                    f"Actual: {r['actual_name'] or '-'} | "
                    f"Exists: {r['exists']} | "
                    f"Status: {r['status']} | "
                    f"Details: {r['details']}\n"
                )
            f.write("\nFOLDER DETAILS\n" + "-" * 100 + "\n")
            for fd in data["folder_data"]:
                tc = (", ".join(f"{k}: {v}" for k, v in fd["file_type_counts"].items())
                      if fd["file_type_counts"] else "None")
                f.write(
                    f"Folder: {fd['directory']} | "
                    f"Depth: {fd['depth']} | "
                    f"Subfolders: {fd['subfolder_count']} | "
                    f"Files: {fd['file_count']} | "
                    f"Size: {self._fmt_size(fd['folder_file_total_size'])} | "
                    f"Modified: {fd['latest_modified']} | "
                    f"Types: {tc} | "
                    f"Status: {fd['status']}\n"
                )
            f.write("\nFILE DETAILS\n" + "-" * 100 + "\n")
            if data["file_details"]:
                for fi in data["file_details"]:
                    f.write(
                        f"Directory: {fi['directory']} | "
                        f"Name: {fi['name']} | "
                        f"Type: {fi['extension']} | "
                        f"Size: {self._fmt_size(fi['size'])} | "
                        f"Modified: {fi['modified']}\n"
                    )
            else:
                f.write("No files found.\n")

    def _apply_sheet_style(self, ws) -> None:
        hf = PatternFill("solid", fgColor=self.C_UCT.lstrip("#"))
        for cell in ws[1]:
            cell.fill      = hf
            cell.font      = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = Border(bottom=Side(style="thin", color="D9D9D9"))
        ws.freeze_panes = "A2"
        for col_cells in ws.columns:
            max_len = max(
                (len(str(cell.value or "")) for cell in col_cells), default=0)
            ws.column_dimensions[
                get_column_letter(col_cells[0].column)].width = min(max_len + 2, 60)

    def _apply_status_fmt(self, ws, col_letter: str) -> None:
        max_row = ws.max_row
        if max_row < 2:
            return
        rng = f"{col_letter}2:{col_letter}{max_row}"
        for status, (bg, fg) in self.STATUS_COLOURS.items():
            ws.conditional_formatting.add(rng, CellIsRule(
                operator="equal",
                formula=[f'"{status}"'],
                fill=PatternFill(start_color=bg, end_color=bg, fill_type="solid"),
                font=Font(color=fg, bold=True),
            ))

    def _write_workbook(self, path: str, data: dict, root: str) -> None:
        wb = Workbook()

        # Summary
        ws = wb.active
        ws.title = "Course Audit Summary"
        for row in [
            ["Item", "Value"],
            ["Root directory",   root],
            ["Auditor",          self.selected_user.get()],
            ["Course TA",        self.selected_ta.get().strip() or "—"],
            ["Detected profile", data["profile"] + ("  [GA course]" if data["is_ga"] else "")],
            ["Scan date",        datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Total folders",    data["total_folders"]],
            ["Total files",      data["total_files"]],
            ["Total size",       self._fmt_size(data["total_size_bytes"])],
            ["Issues found",     len(data["issues"])],
        ]:
            ws.append(row)

        # Expected Structure Check (Status = col F)
        ws2 = wb.create_sheet("Expected Structure Check")
        ws2.append(["Parent Path", "Level", "Expected Name", "Actual Name", "Exists",
                    "Status", "Details", "Reviewer", "Checked", "Comment", "Action Needed"])
        for r in data["expected_results"]:
            ws2.append([r["relative_path"], r["level"], r["expected_name"],
                        r["actual_name"], r["exists"], r["status"], r["details"],
                        "", "", "", ""])

        # Folder Details (Status = col H)
        ws3 = wb.create_sheet("Folder Details")
        ws3.append(["Folder", "Depth", "Subfolder Count", "File Count", "Folder File Size",
                    "Latest Modified", "File Type Counts", "Status",
                    "Reviewer", "Checked", "Comment", "Action Needed"])
        for fd in data["folder_data"]:
            tc = (", ".join(f"{k}: {v}" for k, v in fd["file_type_counts"].items())
                  if fd["file_type_counts"] else "None")
            ws3.append([fd["directory"], fd["depth"], fd["subfolder_count"],
                        fd["file_count"], self._fmt_size(fd["folder_file_total_size"]),
                        fd["latest_modified"], tc, fd["status"], "", "", "", ""])

        # File Details
        ws4 = wb.create_sheet("File Details")
        ws4.append(["Directory", "File Name", "Type", "Size", "Modified"])
        for fi in data["file_details"]:
            ws4.append([fi["directory"], fi["name"], fi["extension"],
                        self._fmt_size(fi["size"]), fi["modified"]])

        # Exceptions (Status = col F)
        ws5 = wb.create_sheet("Exceptions")
        ws5.append(["Parent Path", "Level", "Expected Name", "Actual Name", "Exists",
                    "Status", "Details", "Reviewer", "Checked", "Comment", "Action Needed"])
        for r in data["issues"]:
            ws5.append([r["relative_path"], r["level"], r["expected_name"],
                        r["actual_name"], r["exists"], r["status"], r["details"],
                        "", "", "", ""])

        for ws in (ws, ws2, ws3, ws4, ws5):
            self._apply_sheet_style(ws)
        self._apply_status_fmt(ws2, "F")
        self._apply_status_fmt(ws3, "H")
        self._apply_status_fmt(ws5, "F")
        wb.save(path)

    # ==================================================================== #
    #  Recent directories                                                    #
    # ==================================================================== #

    def _load_recent_dirs(self) -> list[str]:
        if os.path.exists(self.recent_dirs_file):
            try:
                with open(self.recent_dirs_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    if isinstance(data, list):
                        return data
            except (OSError, json.JSONDecodeError):
                pass
        return []

    def _save_recent_dirs(self) -> None:
        try:
            with open(self.recent_dirs_file, "w", encoding="utf-8") as f:
                json.dump(self.recent_directories, f, indent=4)
        except OSError as e:
            self.log(f"Warning: Could not save recent directories: {e}")

    def _update_recent(self, directory: str) -> None:
        if directory in self.recent_directories:
            self.recent_directories.remove(directory)
        self.recent_directories.insert(0, directory)
        self.recent_directories = self.recent_directories[:10]
        self.recent_combo["values"] = self.recent_directories
        self._save_recent_dirs()

    # ── GUI callbacks ──────────────────────────────────────────────────

    def _browse(self) -> None:
        d = filedialog.askdirectory()
        if d:
            self.selected_directory.set(d)
            self._update_recent(d)
            self.log(f"Selected directory: {d}")

    def _select_recent(self, event=None) -> None:
        s = self.recent_combo.get()
        if s:
            self.selected_directory.set(s)

    def _use_recent(self) -> None:
        s = self.recent_combo.get()
        if s:
            self.selected_directory.set(s)
            self.log(f"Selected recent directory: {s}")
        else:
            messagebox.showwarning("No Selection", "Please select a recent directory first.")

    def _clear(self) -> None:
        self.output_text.delete("1.0", tk.END)
        self.tree_text.delete("1.0", tk.END)
        for tv in (self.issues_tv, self.expected_tv, self.folder_tv, self.file_tv):
            for item in tv.get_children():
                tv.delete(item)
        self.summary_label.config(text="No audit run yet.")

    # ==================================================================== #
    #  Main scan entry point                                                 #
    # ==================================================================== #

    def _run_audit(self) -> None:
        root_path = self.selected_directory.get().strip()
        if not root_path:
            messagebox.showwarning("No Directory Selected", "Please select a directory first.")
            return
        if not os.path.isdir(root_path):
            messagebox.showerror("Invalid Directory", "The selected path is not a valid directory.")
            return

        try:
            self._update_recent(root_path)
            self.log("Starting course folder audit...")
            self.log(f"Auditor: {self.selected_user.get()}")

            data = self.analyse(root_path)

            self.log(f"Profile:       {data['profile']}"
                     f"{'  [GA course]' if data['is_ga'] else ''}")
            self.log("Audit complete.")
            self.log(f"Total folders: {data['total_folders']}")
            self.log(f"Total files:   {data['total_files']}")
            self.log(f"Issues found:  {len(data['issues'])}")
            if data["overall_file_types"]:
                self.log("Overall file types:")
                for ext, count in data["overall_file_types"].items():
                    self.log(f"  {ext}: {count}")

            self._populate_all(data)

            stem          = self._output_stem(root_path)
            log_path      = os.path.join(root_path, f"{stem}_folder_audit.txt")
            workbook_path = os.path.join(root_path, f"{stem}_folder_audit.xlsx")

            self._write_log(log_path, data, root_path)
            self._write_workbook(workbook_path, data, root_path)

            self.log(f"Log file created:  {log_path}")
            self.log(f"Workbook created:  {workbook_path}")

            self.summary_label.config(
                text=(f"Profile: {data['profile']}"
                      f"{'  [GA]' if data['is_ga'] else ''}   "
                      f"Folders: {data['total_folders']}   "
                      f"Files: {data['total_files']}   "
                      f"Issues: {len(data['issues'])}"))

            messagebox.showinfo(
                "Success",
                (f"Course folder audit complete.\n\n"
                 f"Profile:  {data['profile']}"
                 f"{'  [GA]' if data['is_ga'] else ''}\n"
                 f"Folders:  {data['total_folders']}\n"
                 f"Files:    {data['total_files']}\n"
                 f"Issues:   {len(data['issues'])}\n\n"
                 f"Log saved to:\n{log_path}\n\n"
                 f"Workbook saved to:\n{workbook_path}"),
            )

        except Exception as exc:
            self.log(f"An error occurred: {exc}")
            self.log(traceback.format_exc())
            messagebox.showerror("Error", f"An error occurred: {exc}")


# ===========================================================================
# Entry point
# ===========================================================================

def main() -> None:
    root = tk.Tk()
    CourseFolderAuditApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
